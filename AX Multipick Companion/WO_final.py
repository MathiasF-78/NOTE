# ============================================================
# IMPORTS
# ============================================================
import os
import threading
import tempfile
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.sql import text, bindparam
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.comments import Comment


# ============================================================
# DATABASE CONNECTION WITH SQLALCHEMY
# ============================================================
engine = create_engine(
    "mssql+pyodbc://DAX-SQL01-PROD.noteone.local/DAX_PROD?driver=SQL+Server&trusted_connection=yes"
)

# ============================================================
# SQL QUERY TO RETREIVE PICKING LISTS FROM WORK ORDER NUMBERS
# ============================================================
def fetch_data(cmpy, work_orders):
  
    query = text("""
    WITH PicklistJournals AS (
        SELECT JOURNALID
        FROM PRODJOURNALTABLE
        WHERE PRODID IN :work_orders
        AND JOURNALTYPE = 0
        AND DATAAREAID = :cmpy
    )

    SELECT 
        pjb.ITEMID AS [ItemId],
        SUM(pjb.BOMCONSUMP) AS [Line QTY],
        a.Total2 AS [Total pick QTY],
        pjb.BOMUNITID AS [Unit],
        id.WMSLOCATIONID AS [Location],
        ISNULL(f.[ON HAND], 0) AS [On hand QTY],
        ISNULL(f.[ON CON PO], 0) AS [QTY On CON/LTB PO],
        CASE WHEN f.[ON HAND] < a.Total2 THEN 'Yes' ELSE '' END AS [Shortage?],
        f.[ON PO] AS [QTY on PO delivery latest today],
        COUNT(DISTINCT pjb.JOURNALID) AS [Number of orders]
    FROM PRODJOURNALBOM AS pjb
    LEFT JOIN INVENTDIM AS id ON pjb.InventDimId = id.InventDimId
    LEFT JOIN (
        SELECT DISTINCT InventSum.ITEMID, D.[ON HAND], C.[ON CON PO], G.[ON PO]
        FROM InventSum
        LEFT JOIN InventDim ON InventSum.INVENTDIMID = InventDim.INVENTDIMID
        LEFT JOIN (
            SELECT InventSum.ITEMID, SUM(InventSum.PHYSICALINVENT) AS [ON HAND]
            FROM InventSum
            INNER JOIN InventDim ON InventSum.INVENTDIMID = InventDim.INVENTDIMID
            WHERE (
                (InventDim.WMSLOCATIONID NOT LIKE 'KAOS%' AND InventSum.DATAAREAID = :cmpy)
                AND InventSum.CLOSED = 0 AND InventSum.PHYSICALINVENT > 0
            ) OR (
                InventSum.CLOSED = 0 AND InventSum.PHYSICALINVENT > 0 AND InventSum.DATAAREAID != :cmpy
            )
            GROUP BY InventSum.ITEMID
        ) D ON D.ITEMID = InventSum.ITEMID
        LEFT JOIN (
            SELECT ITEMID, SUM(REMAINPURCHPHYSICAL) AS [ON CON PO]
            FROM PURCHLINE
            WHERE (PURCHID LIKE 'LTB%' OR PURCHID LIKE 'CON%')
            AND DataAreaId = :cmpy AND IsDeleted = 0
            GROUP BY ITEMID
        ) C ON InventSum.ITEMID = C.ITEMID
        LEFT JOIN (
            SELECT ITEMID, SUM(REMAINPURCHPHYSICAL) AS [ON PO]
            FROM PURCHLINE
            WHERE (PURCHID NOT LIKE 'LTB%' AND PURCHID NOT LIKE 'CON%')
            AND ConfirmedDlv <= GETDATE()
            AND DataAreaId = :cmpy AND IsDeleted = 0
            GROUP BY ITEMID
        ) G ON InventSum.ITEMID = G.ITEMID
        WHERE (
            (InventDim.WMSLOCATIONID NOT LIKE 'KAOS%' AND InventSum.DATAAREAID = :cmpy)
            AND InventSum.CLOSED = 0 AND InventSum.POSTEDQTY > 0
        ) OR (
            InventSum.CLOSED = 0 AND InventSum.POSTEDQTY > 0 AND InventSum.DATAAREAID != :cmpy
        )
    ) f ON f.ITEMID = pjb.ItemId
    LEFT JOIN (
        SELECT ItemId, SUM(BOMCONSUMP) AS Total2
        FROM PRODJOURNALBOM
        WHERE JOURNALID IN (SELECT JOURNALID FROM PicklistJournals)
        AND DataAreaId = :cmpy
        GROUP BY ItemId
    ) a ON a.ItemId = pjb.ItemId
    WHERE pjb.JOURNALID IN (SELECT JOURNALID FROM PicklistJournals)
    AND pjb.DataAreaId = :cmpy
    GROUP BY 
        pjb.ITEMID, pjb.BOMUNITID, id.WMSLOCATIONID,
        f.[ON HAND], f.[ON CON PO], a.Total2, f.[ON PO]
    ORDER BY pjb.ItemId;
    """).bindparams(bindparam("work_orders", expanding=True), bindparam("cmpy"))

# Kör fråga och returnera dataframe:
    df = pd.read_sql(query, engine, params={"cmpy": "se04", "work_orders": work_orders})
    return df

# ============================================================
# STATUSLOGGER 
# ============================================================
class StatusLogger:
    def __init__(self, parent):
        self.label = tk.Label(parent, text="", anchor="w", justify="left", font=("Segoe UI", 9), height=2, wraplength=500)
        self.label.grid(row=99, column=0, columnspan=10, sticky="we", padx=5, pady=2)
        self.current_message = ""
        self.dot_count = 0
        self.running = False

    def start(self, message):
        self.current_message = message
        self.dot_count = 0
        self.running = True
        self.update()

    def update(self):
        if self.running:
            dots = "." * (self.dot_count % 4)
            self.label.config(text=f"{self.current_message}{dots}")
            self.dot_count += 1
            self.label.after(500, self.update)

    def stop(self, final_message=None):
        self.running = False
        if final_message:
            self.label.config(text=final_message)

# ============================================================
# FORMATTING FOR OUTPUT
# ============================================================
def create_formatted_excel(df):
    if df.empty:
        return None

    # Anpassade rubriker med radbrytningar
    custom_headers = {
        "Line QTY": "QTY",
        "Total pick QTY": "Total\nQTY",
        "On hand QTY": "In\nStock",
        "QTY On CON/LTB PO": "CON/LTB\nPO QTY",
        "QTY on PO delivery latest today": "ETD\nToday",
        "Number of orders": "#\norders"
    }
    df.columns = [custom_headers.get(col, col) for col in df.columns]

    wb = Workbook()
    ws = wb.active
    ws.title = "Result"
    
    #print(df.columns.tolist())

    # Fasta kolumnbredder
    fixed_widths = {
        "A": 16,   # ItemID (tidigare 15)
        "B": 9,    # QTY (tidigare 8)
        "C": 9,    # Total QTY
        "D": 5.5,  # Unit (tidigare 4.5)
        "E": 9,    # Location
        "F": 9,    # In Stock
        "G": 10,   # CON/LTB PO QTY
        "H": 9,    # Shortage?
        "I": 8,    # ETD Today
        "J": 8     # # orders
    }



    for col_letter, width in fixed_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Färger och stilar
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alt_fill = PatternFill(start_color="D8E0F3", end_color="D8E0F3", fill_type="solid")
    ws.row_dimensions[1].height = 42

    # Skriv in data och formatera
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                # Anpassad fontstorlek för G1, H1, I1
                if c_idx in [7, 8, 9]:  # G=7, H=8, I=9
                    cell.font = Font(color="FFFFFF", bold=True, size=9)    
                    # Lägg till kommentarer
                    comments = {
                        7: "QTY On CON/LTB PO",
                        8: "Shortage?",
                        9: "QTY on PO delivery latest today"
                    }
                    cell.comment = Comment(comments[c_idx], "System")

                else:
                    cell.font = header_font
            elif r_idx % 2 == 0:
                cell.fill = alt_fill


    # Lås rubrikraden
    ws.freeze_panes = "A2"

    # Anpassa kolumnbredd till innehåll
    min_width = 15
    max_width = 40
    
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        if column_letter in fixed_widths:
            continue
        max_length = 0
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max(min_width, min(max_length + 2, max_width))
        ws.column_dimensions[column_letter].width = adjusted_width

    # Sidinställningar för A4 liggande
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    # Spara tillfällig fil
    import tempfile
    import os
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        file_path = tmp.name
        wb.save(file_path)

    return file_path

# ============================================================
# PRINT DIRECT FROM APP WINDOW
# ============================================================
def print_excel(df):
    if df.empty:
        messagebox.showinfo("Ingen data", "Det finns ingen data att skriva ut.")
        return

    try:
        file_path = create_formatted_excel(df)
        if not file_path:
            messagebox.showerror("Fel", "Kunde inte skapa Excel-filen för utskrift.")
            return

        if os.name == 'nt':    
            # === TESTLÄGE: öppna istället för att skriva ut ===
            # os.startfile(file_path)  # ← Kommentera bort denna rad för att aktivera utskrift
            os.startfile(file_path, "print")  # ← Avkommentera denna rad för att skriva ut direkt
        else:
            os.system(f"lp {file_path}")

    except Exception as e:
        messagebox.showerror("Fel", f"Kunde inte skriva ut Excel-filen:\n{e}")

# ============================================================
# GUI APPLICATION (TKINTER)
# ============================================================
class JournalApp:
    # ============================================================
    # INIT
    # ============================================================
    def __init__(self, root):
        self.root = root
        self.root.title("AX Multipick Companion")
        self.status_logger = StatusLogger(root)
        self.cancel_requested = False # Flagga för att avbryta
        self.df = pd.DataFrame()

        tk.Label(root, text="Company Code:").grid(row=0, column=0, sticky="w")

        self.cmpy_entry = tk.Entry(root, width=20)
        self.cmpy_entry.grid(row=0, column=1, columnspan=2, sticky="we")
        self.cmpy_entry.insert(0, "SE04")
        self.cmpy_entry.config(state="readonly", readonlybackground="#f0f0f0", fg="black")

        self.work_order_entries = []
        for i in range(10):
            tk.Label(root, text=f"Work Order ID {i+1}:").grid(row=i+1, column=0, sticky="w")
            entry = tk.Entry(root)
            entry.grid(row=i+1, column=1, columnspan=2, sticky="we")
            self.work_order_entries.append(entry)

        self.fetch_button = tk.Button(root, text="Fetch Data", command=self.fetch_and_display)
        self.fetch_button.grid(row=11, column=0, columnspan=3, pady=10)

        self.tree = ttk.Treeview(root, show="headings")
        self.tree.grid(row=12, column=0, columnspan=3, sticky="nsew")

# ============================================================
# STYLE FORMATTING FOR TABLE (TREEVIEW - TKINTER)
# ============================================================
        style = ttk.Style()
        style.configure("Treeview",
                        background="white",
                        foreground="black",
                        rowheight=25,
                        fieldbackground="white",
                        borderwidth=1,
                        relief="solid")

        style.configure("Treeview.Heading",
                        font=("Helvetica", 10, "bold"),
                        borderwidth=1,
                        relief="solid")

        # Alternativa rader för bättre läsbarhet
        self.tree.tag_configure("evenrow", background="#f2f2f2")
        self.tree.tag_configure("oddrow", background="#ffffff")

        scrollbar = ttk.Scrollbar(root, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.config(height=20)
        scrollbar.grid(row=12, column=3, sticky="ns")

        root.grid_rowconfigure(12, weight=1)
        root.grid_rowconfigure(13, weight=0)

        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        root.maxsize(screen_width, screen_height)

# ============================================================
# BUTTONS (TKINTER)
# ============================================================               
        self.export_button = tk.Button(root, text="📊 Exportera till Excel", command=lambda: self.export_and_open_excel_with_formatting_and_fit(self.df, self.status_logger))
        self.export_button.grid(row=13, column=0, sticky="w", pady=10)

        self.root.update()
        self.root.minsize(self.root.winfo_width(), self.root.winfo_height())

        self.print_button = tk.Button(root, text="🖨️ Skriv ut", command=lambda: print_excel(self.df))
        self.print_button.grid(row=13, column=1, sticky="w", pady=10)

        self.cancel_button = tk.Button(root, text="❌ Avbryt", command=self.cancel_operation)
        self.cancel_button.grid(row=13, column=2, sticky="e", pady=10)

    # ============================================================
    # MENU BAR
    # ============================================================
        menubar = tk.Menu(self.root)

        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Inställningar", command=self.open_settings)
        file_menu.add_separator()
        file_menu.add_command(label="Avsluta", command=self.root.quit)
        menubar.add_cascade(label="Arkiv", menu=file_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="Om", command=self.show_about)
        menubar.add_cascade(label="Hjälp", menu=help_menu)

        self.root.config(menu=menubar)

    # ============================================================
    # ABOUT MENU
    # ============================================================
    def show_about(self):
        messagebox.showinfo(
            "Om",
            "AX Multipick Companion\n\n"
            "Version: 1.0.0\n"
            "Utvecklad av: Mathias Fredriksson\n\n"
            "Detta verktyg är skapat för att förenkla hantering av multipick-data\n"
            "från AX-systemet. Programmet möjliggör inmatning av arbetsorder,\n"
            "hämtning av data, export till Excel samt utskrift.\n\n"
        )

    # ============================================================
    # SETTINGS MENU
    # ============================================================
    def open_settings(self):
        # Skapa nytt fönster
        settings_window = tk.Toplevel(self.root)
        settings_window.title("Inställningar")
        settings_window.geometry("300x150")
        settings_window.grab_set()  # Gör fönstret modal

        tk.Label(settings_window, text="Välj Company Code:").pack(pady=10)

        # Alternativ för rullistan
        company_codes = ["SE04", "DK01", "FI02", "NO03", "DE05"]

        # Skapa en StringVar för att hålla det valda värdet
        selected_code = tk.StringVar()
        selected_code.set(self.cmpy_entry.get())  # Förvalt värde

        # Skapa rullistan
        dropdown = ttk.Combobox(settings_window, textvariable=selected_code, values=company_codes, state="readonly")
        dropdown.pack(pady=5)

        # Bekräfta-knapp
        def apply_selection():
            self.cmpy_entry.config(state="normal")
            self.cmpy_entry.delete(0, tk.END)
            self.cmpy_entry.insert(0, selected_code.get())
            self.cmpy_entry.config(state="readonly")
            settings_window.destroy()

        tk.Button(settings_window, text="OK", command=apply_selection).pack(pady=10)

    
# ============================================================
# CANCEL FUNCTION
# ============================================================
    def cancel_operation(self):
        self.cancel_requested = True
        if self.status_logger:
            self.status_logger.stop("Avbrutet av användaren ❌")

# ============================================================
# INITIATE RETRIEVAL AND START LOG ANIMATION
# ============================================================
    def fetch_and_display(self):
        self.cancel_requested = False
        if self.status_logger:
            self.status_logger.start("Hämtar data...")
        threading.Thread(target=self._fetch_and_display_data).start()

# ============================================================
# RETREIVE DATA AND DISPLAY IN TABLE
# ============================================================
    def _fetch_and_display_data(self):
        if self.cancel_requested:
            return

        cmpy = self.cmpy_entry.get().strip()
        work_orders = [entry.get().strip() for entry in self.work_order_entries if entry.get().strip()]

        if not cmpy or not work_orders:
            if self.status_logger:
                self.status_logger.stop("Fel: saknar input ❌")
            messagebox.showerror("Input Error", "Please enter company code and at least one Work Order ID.")
            return

        try:
            df = fetch_data(cmpy, work_orders)

            if df.empty:
                if self.status_logger:
                    self.status_logger.stop("Inga rader hittades")
                messagebox.showinfo("No Data", "No records returned from the database.")
                return

            self.df = df  # Spara för export

            # Rensa och bygg om Treeview
            self.tree.delete(*self.tree.get_children())
            self.tree["columns"] = list(df.columns)

            for col in df.columns:
                self.tree.heading(col, text=col)
                self.tree.column(col, width=120, anchor="center")

            for i, (_, row) in enumerate(df.iterrows()):
                tag = "evenrow" if i % 2 == 0 else "oddrow"
                self.tree.insert("", "end", values=list(row), tags=(tag,))

            # Justera höjd på Treeview
            num_rows = len(df)
            max_rows = 30
            visible_rows = min(num_rows, max_rows)
            self.tree.config(height=visible_rows)

            # Justera fönsterhöjd
            self.root.update_idletasks()
            screen_height = self.root.winfo_screenheight()
            current_width = self.root.winfo_width()
            current_height = self.root.winfo_height()
            max_height = screen_height - 100
            new_height = min(current_height, max_height)
            self.root.geometry(f"{current_width}x{new_height}+0+0")

            if self.status_logger:
                self.status_logger.stop("Data hämtad ✅")

        except Exception as e:
            if self.status_logger:
                self.status_logger.stop("Fel vid hämtning ❌")
            messagebox.showerror("Error", str(e))

# ============================================================
# EXPORT TO EXCEL AND OPEN
# ============================================================
    def export_and_open_excel_with_formatting_and_fit(self, df, status_logger):
        if self.status_logger:
            self.status_logger.start("Förbereder export")

        if df.empty:
            messagebox.showinfo("Ingen data", "Det finns ingen data att exportera.")
            return

        try:
            file_path = create_formatted_excel(df)
            if not file_path:
                raise Exception("Kunde inte skapa Excel-filen.")

            if self.status_logger:
                self.status_logger.stop("Export klar ✅")

            # Öppna filen för inspektion
            if os.name == 'nt':
                os.startfile(file_path)
            else:
                os.system(f"open {file_path}")

        except Exception as e:
            messagebox.showerror("Fel", f"Kunde inte exportera Excel-filen:\n{e}")
            if status_logger:
                status_logger.stop("Fel vid export ❌")


# ============================================================
# ENTRY POINT – START THE APPLICATION AND HANDLE EXIT
# ============================================================
if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = JournalApp(root)
        root.mainloop()
    except KeyboardInterrupt:
        print("Programmet avbröts av användaren.")
        try:
            root.destroy()
        except:
            pass
